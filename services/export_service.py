"""
Serviço de Exportação (Excel).
"""

import os
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

class ExportService:
    @staticmethod
    def export_to_excel(title: str, columns: list, data: list, default_dir: str) -> str:
        """
        Exporta dados para uma planilha Excel.
        Retorna o caminho do arquivo gerado ou lança exceção.
        columns: Lista de nomes de colunas (labels).
        data: Lista de listas contendo os dados, na ordem das colunas.
        """
        if not openpyxl:
            raise ImportError("A biblioteca openpyxl não está instalada. Contate o suporte.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatorio"

        # Título
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Espaçamento
        ws.append([])

        # Cabeçalhos
        ws.append(columns)
        header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
        header_font = Font(color="ffffff", bold=True)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        # Dados
        for item in data:
            ws.append(item)

        # Ajustar larguras de coluna
        from openpyxl.utils import get_column_letter
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in col:
                try: 
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

        filename = f"Relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(default_dir, filename)
        
        wb.save(filepath)
        return filepath
